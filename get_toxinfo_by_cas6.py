import aiohttp
import asyncio
import json
from typing import List, Dict, Any
import sys
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

BASE_URL = "https://pubchem.ncbi.nlm.nih.gov/rest/pug"
VIEW_URL = "https://pubchem.ncbi.nlm.nih.gov/rest/pug_view"

async def fetch_url(session: aiohttp.ClientSession, url: str, retries: int = 3) -> Dict[str, Any]:
    for attempt in range(retries):
        try:
            async with session.get(url) as response:
                if response.status == 200:
                    try:
                        # Try to decode as UTF-8 first
                        text = await response.text()
                        return json.loads(text)
                    except UnicodeDecodeError:
                        # If UTF-8 decoding fails, use a more permissive encoding
                        content = await response.read()
                        text = content.decode('iso-8859-1', errors='ignore')
                        return json.loads(text)
                elif response.status == 503:
                    await asyncio.sleep(2 ** attempt)
                else:
                    print(f"Error {response.status} for URL: {url}")
                    return {'error': f"HTTP error {response.status}"}
        except Exception as e:
            print(f"Exception for URL {url}: {str(e)}")
            if attempt == retries - 1:
                return {'error': str(e)}
            await asyncio.sleep(2 ** attempt)

async def get_pubchem_cid(session: aiohttp.ClientSession, cas_number: str) -> int:
    url = f"{BASE_URL}/compound/name/{quote(cas_number)}/cids/JSON"
    data = await fetch_url(session, url)
    if 'IdentifierList' in data and 'CID' in data['IdentifierList']:
        return data['IdentifierList']['CID'][0]
    print(f"Failed to get CID for CAS {cas_number}. Response: {data}")
    return None

async def get_compound_data(session: aiohttp.ClientSession, cid: int) -> Dict[str, Any]:
    url = f"{VIEW_URL}/data/compound/{cid}/JSON"
    return await fetch_url(session, url)

async def get_iupac_and_smiles(session: aiohttp.ClientSession, cid: int) -> Dict[str, str]:
    url = f"{BASE_URL}/compound/cid/{cid}/property/IUPACName,CanonicalSMILES/JSON"
    data = await fetch_url(session, url)
    if 'PropertyTable' in data and 'Properties' in data['PropertyTable']:
        properties = data['PropertyTable']['Properties'][0]
        return {
            'IUPAC': properties.get('IUPACName', 'N/A'),
            'SMILES': properties.get('CanonicalSMILES', 'N/A')
        }
    return {'IUPAC': 'N/A', 'SMILES': 'N/A'}

async def get_names_and_synonyms(session: aiohttp.ClientSession, cid: int) -> Dict[str, Any]:
    url = f"{VIEW_URL}/data/compound/{cid}/JSON"
    data = await fetch_url(session, url)
    names_and_synonyms = {'Names': [], 'Synonyms': []}
    if 'Record' in data and 'Section' in data['Record']:
        for section in data['Record']['Section']:
            if section.get('TOCHeading') == 'Names and Identifiers':
                for subsection in section.get('Section', []):
                    if subsection.get('TOCHeading') == 'Computed Descriptors':
                        for info in subsection.get('Information', []):
                            if info.get('Name') == 'IUPAC Name':
                                names_and_synonyms['Names'].append(info['Value']['StringWithMarkup'][0]['String'])
                    elif subsection.get('TOCHeading') == 'Synonyms':
                        for info in subsection.get('Information', []):
                            if info.get('Name') == 'Synonym':
                                names_and_synonyms['Synonyms'].extend([value['String'] for value in info['Value']['StringWithMarkup']])
                break
    return names_and_synonyms

async def get_literature_references(session: aiohttp.ClientSession, cid: int) -> Dict[str, List[str]]:
    url = f"{VIEW_URL}/data/compound/{cid}/JSON"
    data = await fetch_url(session, url)
    references = {
        'Disease and References': [],
        'Nature Journal References': [],
        'Springer Nature References': [],
        'Other Safety Information': []
    }
    if 'Record' in data and 'Section' in data['Record']:
        for section in data['Record']['Section']:
            if section.get('TOCHeading') in references:
                for info in section.get('Information', []):
                    if 'Value' in info and 'StringWithMarkup' in info['Value']:
                        references[section['TOCHeading']].extend([item['String'] for item in info['Value']['StringWithMarkup']])
    return references

def extract_tox_data(data: Dict[str, Any]) -> Dict[str, Any]:
    tox_data = {}
    if 'Record' not in data or 'Section' not in data['Record']:
        return tox_data

    tox_keywords = ['tox', 'safety', 'hazard', 'health', 'exposure', 'risk', 'carcinogen']
    
    def process_section(section):
        if 'TOCHeading' in section and any(keyword in section['TOCHeading'].lower() for keyword in tox_keywords):
            heading = section['TOCHeading']
            tox_data[heading] = []
            if 'Information' in section:
                for info in section['Information']:
                    if 'Value' in info and 'StringWithMarkup' in info['Value']:
                        tox_data[heading].extend([item['String'] for item in info['Value']['StringWithMarkup']])
            if 'Section' in section:
                for subsection in section['Section']:
                    process_section(subsection)

    for section in data['Record']['Section']:
        process_section(section)

    return tox_data

async def process_cas_number(session: aiohttp.ClientSession, cas_number: str) -> Dict[str, Any]:
    print(f"Processing CAS {cas_number}")
    try:
        cid = await get_pubchem_cid(session, cas_number)
        if cid is None:
            return {'CAS': cas_number, 'error': 'Could not find PubChem CID'}
        
        print(f"Found CID {cid} for CAS {cas_number}")
        compound_data = await get_compound_data(session, cid)
        if 'error' in compound_data:
            return {'CAS': cas_number, 'PubChemCID': cid, 'error': compound_data['error']}
        
        iupac_smiles = await get_iupac_and_smiles(session, cid)
        names_and_synonyms = await get_names_and_synonyms(session, cid)
        literature_references = await get_literature_references(session, cid)
        
        if 'Record' not in compound_data:
            return {
                'CAS': cas_number,
                'PubChemCID': cid,
                'IUPAC': iupac_smiles['IUPAC'],
                'SMILES': iupac_smiles['SMILES'],
                'Names': names_and_synonyms['Names'],
                'Synonyms': names_and_synonyms['Synonyms'],
                'LiteratureReferences': literature_references,
                'error': 'Could not find compound data'
            }
        
        tox_data = extract_tox_data(compound_data)
        return {
            'CAS': cas_number,
            'PubChemCID': cid,
            'IUPAC': iupac_smiles['IUPAC'],
            'SMILES': iupac_smiles['SMILES'],
            'Names': names_and_synonyms['Names'],
            'Synonyms': names_and_synonyms['Synonyms'],
            'LiteratureReferences': literature_references,
            'ToxData': tox_data
        }
    except Exception as e:
        print(f"Error processing CAS {cas_number}: {str(e)}")
        return {'CAS': cas_number, 'error': f"Processing error: {str(e)}"}

async def get_tox_data_for_cas_numbers(cas_numbers: List[str]) -> List[Dict[str, Any]]:
    async with aiohttp.ClientSession() as session:
        tasks = [process_cas_number(session, cas) for cas in cas_numbers]
        return await asyncio.gather(*tasks)

def save_to_json(data: List[Dict[str, Any]], filename: str):
    with open(filename, 'w') as f:
        json.dump(data, f, indent=2)

def save_to_excel(data: List[Dict[str, Any]], filename: str):
    wb = Workbook()
    wb.remove(wb.active)
    bold_font = Font(bold=True)

    for compound in data:
        cas = compound['CAS']
        ws = wb.create_sheet(title=cas)
        
        ws['A1'], ws['B1'] = 'CAS Number', cas
        ws['A2'], ws['B2'] = 'PubChem CID', compound.get('PubChemCID', 'N/A')
        ws['A3'], ws['B3'] = 'IUPAC Name', compound.get('IUPAC', 'N/A')
        ws['A4'], ws['B4'] = 'SMILES', compound.get('SMILES', 'N/A')

        row = 6
        ws.cell(row=row, column=1, value='Names').font = bold_font
        row += 1
        for name in compound.get('Names', []):
            ws.cell(row=row, column=2, value=name)
            row += 1
        row += 1

        ws.cell(row=row, column=1, value='Synonyms').font = bold_font
        row += 1
        for synonym in compound.get('Synonyms', []):
            ws.cell(row=row, column=2, value=synonym)
            row += 1
        row += 1

        ws.cell(row=row, column=1, value='Literature References').font = bold_font
        row += 1
        for ref_type, references in compound.get('LiteratureReferences', {}).items():
            ws.cell(row=row, column=1, value=ref_type).font = bold_font
            row += 1
            for ref in references:
                ws.cell(row=row, column=2, value=ref)
                row += 1
            row += 1

        ws.cell(row=row, column=1, value='Toxicological Data').font = bold_font
        row += 1

        if 'error' in compound:
            ws.cell(row=row, column=1, value=f"Error: {compound['error']}")
        elif 'ToxData' in compound:
            for category, data in compound['ToxData'].items():
                ws.cell(row=row, column=1, value=category).font = bold_font
                row += 1
                for item in data:
                    ws.cell(row=row, column=2, value=item)
                    row += 1
                row += 1
        else:
            ws.cell(row=row, column=1, value="No toxicological data found")

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filename)
    print(f"Data exported to Excel file: {filename}")

def get_cas_numbers_from_user() -> List[str]:
    print("Enter CAS numbers (one per line). Press Enter twice to finish:")
    cas_numbers = []
    while True:
        cas = input().strip()
        if cas:
            cas_numbers.append(cas)
        else:
            break
    return cas_numbers

if __name__ == "__main__":
    if len(sys.argv) > 1:
        cas_numbers = sys.argv[1:]
    else:
        cas_numbers = get_cas_numbers_from_user()

    if not cas_numbers:
        print("No CAS numbers provided. Exiting.")
        sys.exit(1)

    print(f"Fetching data for {len(cas_numbers)} CAS numbers...")
    results = asyncio.run(get_tox_data_for_cas_numbers(cas_numbers))
    
    json_filename = "tox_data.json"
    save_to_json(results, json_filename)
    print(f"Data for {len(results)} compounds saved to {json_filename}")

    excel_filename = "tox_data.xlsx"
    save_to_excel(results, excel_filename)

    print("Please check the JSON and Excel files for the retrieved data.")
